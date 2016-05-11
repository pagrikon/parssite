# Copyright: 2016, Pavel Konurkin
# Author: Pavel Konurkin
# License: BSD
"""
Library for parsing sites
"""
from __future__ import print_function
from grab import Grab
from lxml import etree
import unicodedata
import re
from enum import Enum
import urlparse
import urllib
import copy
import pickle
import os
import os.path
import time
import random
import grab.error
import datetime
import sys
import uuid
import hashlib
import glob


class ParsException(Exception):

    def __init__(self, *args, **kwargs):
        Exception.__init__(self, *args)
        self._attrSet = set()
        for attr in kwargs:
            self[attr] = kwargs[attr]

    def __getattr__(self, name):
        return None

    def __iter__(self):
        return iter(self._attrSet)

    def __getitem__(self, index):
        return getattr(self, index)

    def __setitem__(self, index, value):
        setattr(self, index, value)
        self._attrSet.add(index)

    def __str__(self):
        result = Exception.__str__(self)
        if len(result) > 0:
            result = '{' + result + '}'
        exceptionNames = set()
        for attrName in self:
            value = self[attrName]
            if type(value) is unicode:
                value = value.encode(ParsBase._encoding)
            if isinstance(value, Exception):
                exceptionNames.add(attrName)
            else:
                if attrName != 'page':
                    result += '{' + attrName + ': ' + str(value) + '}'
        for attrName in exceptionNames:
            value = self[attrName]
            result += '{Exception ' + className(value) + ': ' + str(value) + '}'
        result = delElements(result, '\r\n')
        result = normalizeSpace(result)
        return result


class ParsError(ParsException):
    pass

    # def __init__(self, *args, **kwargs):
        # self.errorObj = kwargs.pop('errorObj', None)
        # Exception.__init__(self, *args, **kwargs)


class DuplicateTree(ParsException):
    pass


class StopParsing(ParsException):
    pass


class ParsFileSystemError(ParsError):
    pass


class StructureError(ParsError):
    pass


class DirectionError(ParsError):
    pass


class ControlResultError(ParsError):
    pass


class DuplicateChildName(ParsError):
    pass


class UndefinedInstance(ParsError):
    pass


class UndefinedParent(ParsError):
    pass


class UndefinedQuery(ParsError):
    pass


class BadQueryResult(ParsError):
    pass

    # def __init__(self, *args, **kwargs):
        # self.queryResult = kwargs.pop('queryResult', None)
        # ParsError.__init__(self, *args, **kwargs)


class SetChildError(ParsError):
    pass


class UndefinedChildName(ParsError):
    pass


class UndefinedQueryRootName(ParsError):
    pass


class WebError(ParsError):
    pass


class BadPageError(WebError):
    pass


class WebInternalError(WebError):
    pass


class PageControlFault(WebError):
    pass


class WebServerError(WebError):
    pass


class WebClientError(WebError):
    pass


class ProxyError(WebError):
    pass


class ProxyServerError(ProxyError):
    pass


class AllProxyAlreadyUsed(ProxyError):
    pass


class PageCacheError(ParsError):
    pass


class PageCacheWarning(ParsException):
    pass


class ParsBreak(ParsException):
    pass


class PageError(ParsError):
    pass


class BreakerError(ParsError):
    pass


class _Breaker():

    def __init__(self, maxIterNumber=None):
        self.iterNumber = 0
        self.maxIterNumber = maxIterNumber

    def __call__(self, maxIterNumber=None):
        self.iterNumber += 1
        if maxIterNumber is None:
            maxIterNumber = self.maxIterNumber
        if maxIterNumber is None:
            raise BreakerError('maxIterNumber is not defined')
        if self.iterNumber > maxIterNumber:
            raise ParsBreak


_breaker = _Breaker()

def md5(value):
    return hashlib.md5(value).hexdigest()

def mkdirs(path):
    if os.path.exists(path):
        if not os.path.isdir(path):
            raise ParsFileSystemError('Cannot create dir: '+path)
    else:
        os.makedirs(path)

def normalizePath(path, itDir=False):
    path = path.strip()
    path = path.replace('//','/')
    if itDir and path[-1] != '/':
        path += '/'
    return path

def splitDirFile(path):
    pathList = path.rsplit('/', 1)
    dirs = pathList[0] + '/'
    fileName = pathList[1]
    return dirs, fileName

def write(data, path, mode='w'):
    if type(path) is tuple:
        dirs, fileName = path
        dirs = normalizePath(dirs, itDir=True)
        path = dirs + fileName
        path = normalizePath(path)
    else:
        path = normalizePath(path)
        dirs, fileName = splitDirFile(path)
    mkdirs(dirs)
    outFile = open(path, mode)
    outFile.write(data)
    outFile.close()

def className(instance, withPath=True):
    name = str(instance.__class__)
    if name[0] == '<':
        name = name[8:-2]
    if not withPath:
        name = name.rsplit('.', 1)[1]
    return name

def isParsStructure(obj):
    if isinstance(obj, ParsBase):
        return True
    elif type(obj) is list:
        for elem in obj:
            if not isParsStructure(elem):
                return False
    elif type(obj) is dict:
        for key in obj:
            if not isParsStructure(obj[key]):
                return False
    else:
        return False
    return True

def normalizeUrl(url):
    resultUrl = url
    # ------------------------------
    # print(type(url))
    # # try:
        # # print(url)
    # # except Exception:
        # # pass
    # if type(url) is str or type(url) is Url:
        # print(url)
    # else:
        # print(unicode(url).encode('utf-8'))
    # -----------------------------
    resultUrl = urllib.unquote(resultUrl)
    resultUrl = resultUrl.strip()
    return resultUrl

def normalizeSpace(string):
    if type(string) is unicode:
        result = u''
        spaceSymbol = u' '
    else:
        result = ''
        spaceSymbol = ' '
    firstSpace = True
    for ch in string:
        if ch.isspace():
            if firstSpace:
                firstSpace = False
                result += spaceSymbol
        else:
            firstSpace = True
            result += ch
    result = result.strip()
    return result

def createDateFileName():
    return datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')

def createDatePidFileName():
    return createDateFileName() + '.' + str(os.getpid())



class Direction(Enum):
    forward = 1
    backward = 2


class Structure(Enum):
    single = 1
    list = 2
    dict = 3


class FilePathType(Enum):
    urlPath = 0
    uuidSingleDir = 1
    uuidMultiDir = 2


class ControlResult(Enum):
    fault = 0
    bad = 1
    ok = 2


def printUnicodeInfoByChr(str):
    for ch in str:
        info = "ch='" + ch.encode(ParsBase._encoding) + "' " + \
            "category='" + unicodedata.category(ch) + "' " + \
            "escape='" + ch.encode('unicode_escape') + "' " + \
            "name='" + unicodedata.name(ch, 'UNDEFINE_NAME') + "'"
        print(info)


def delElements(sequence, delSequence):
    result = sequence[0: 0]
    for elem in sequence:
        insert = True
        for delElem in delSequence:
            if elem == delElem:
                insert = False
                break
        if insert:
            result += elem
    return result


class DelUnicodeSymbols(object):

    def __init__(self, symbols):
        self.symbols = symbols

    def __call__(self, unistr):
        if len(self.symbols) == 0:
            return unistr
        else:
            outstr = u''
            for ch in unistr:
                for delch in self.symbols:
                    if ch == delch:
                        ch = ''
                        break
                outstr += ch
            return outstr


class ReplaceUnicodeSubStrings(object):

    def __init__(self, replaceList):
        self.replaceList = replaceList

    def __call__(self, unistr):
        if len(self.replaceList) == 0:
            return unistr
        else:
            outstr = unistr
            for (old, new) in self.replaceList:
                outstr = outstr.replace(old, new)
            return outstr


class RegexCache(object):

    _cache = {}

    @staticmethod
    def compile(regex):
        pattern = RegexCache._cache.get(regex, None)
        if pattern is None:
            pattern = re.compile(regex, re.U)
            RegexCache._cache[regex] = pattern
        return pattern


class Query:

    def __init__(self, *args, **kwargs):
        self.args = args
        self.kwargs = kwargs
        # self.queryName = '_' + str(self.__class__).rsplit('.', 1)[1]
        self.queryName = '_' + className(self, withPath=False)
        self.queryResultProcessingName = self.queryName + '_'

    def getTargetInstance(self, instance):
        targetInstance = instance.parent
        if targetInstance is None:
            raise UndefinedParent
        return targetInstance

    def __call__(self, instance, *args, **kwargs):
        _args = self.args + args
        _kwargs = self.kwargs.copy()
        for key in kwargs:
            _kwargs[key] = kwargs[key]
        targetInstance = self.getTargetInstance(instance)
        query = getattr(targetInstance, self.queryName)
        queryResult = query(*_args, **_kwargs)
        queryResultProcessing = getattr(instance
                                        , self.queryResultProcessingName, None)
        if queryResultProcessing is not None:
            queryResult = queryResultProcessing(queryResult)
        return queryResult


class value(Query):

    def getTargetInstance(self, instance):
        return self

    def _value(self, value):
        return [value]


class valueList(Query):

    def getTargetInstance(self, instance):
        return self

    def _valueList(self, value):
        return value


class xpath(Query):
    pass


class href(Query):
    pass


class title(Query):
    pass


class regex(Query):
    pass


class regexConcat(Query):
    pass


class reValue(Query):
    pass


class tail(Query):
    pass


class text(Query):
    pass


class area(Query):
    pass


class reValueUnit(Query):
    pass


class number(Query):
    pass


class strip(Query):
    pass


class selfUrl(Query):
    pass


class XpathQueryMixin(object):

    def _xpath(self, xpath):
        return self._elem.xpath(xpath)

    def _attribute(self, attribute, xpath=None):
        if xpath is None or xpath == '':
            xpath = './/'
        if xpath[len(xpath)-1] != '/':
            xpath += '/'
        xpath += 'attribute::'+attribute
        return self._elem.xpath(xpath)

    def _href(self, xpath=None):
        return self._attribute('href', xpath)

    def _title(self, xpath=None):
        return self._attribute('title', xpath)

    def _text(self, text, xpath=None):
        text = normalizeSpace(text)
        if xpath is None or xpath == '':
            xpath = './/'
        if xpath[len(xpath)-1] == '/':
            xpath += '*'
        xpath += "[normalize-space(text())='"+text+"']"
        return self._elem.xpath(xpath)

    def _tail(self, xpath=None):
        result = []
        if xpath is None:
            result.append(self._elem.tail)
        else:
            elems = self._xpath(xpath)
            for elem in elems:
                result.append(elem.tail)
        return result

    def _area(self, xpath, predicate):
        elements = self._xpath(xpath)
        result = []
        predicate = u'self::*' + predicate
        for element in elements:
            if len(element.xpath(predicate)) == 0:
                break
            result.append(element)
        return result


class RegexQueryMixin(object):

    def _regex(self, regex):
        pattern = RegexCache.compile(regex)
        return pattern.findall(self._unicode)

    def _regexConcat(self, regex):
        pattern = RegexCache.compile(regex)
        l = pattern.findall(self._unicode)
        result = []
        for elem in l:
            stri = u''
            if type(elem) is tuple:
                for e in elem:
                    if type(e) is not unicode:
                        e = unicode(e, ParsBase._encoding)
                    stri += e
            else:
                if type(elem) is not unicode:
                    elem = unicode(elem, ParsBase._encoding)
                stri = elem
            result.append(stri)
        return result

    def _number(self):
        digitRegex = r'([\d.,]+)'
        return self._regex(digitRegex)

    def _strip(self):
        return [unicode(self).strip()]

    def _reValue(self, unit, direction=Direction.forward):
        digitRegex = r'([\d.,]+)'
        if direction == Direction.forward:
            regex = digitRegex + r'\s*' + unit
        elif direction == Direction.backward:
            regex = unit + r'\s*' + digitRegex
        else:
            raise DirectionError('Direction should by Direction.forward or' \
                                 + ' Direction.backward')
        return self._regex(regex)

    def _reValueUnit(self, *args):
        # Arguments processing
        if len(args) > 2:
            raise TypeError('_reValueUnit takes at most 3 arguments ('\
                            + len(args)+1 + ' given)')
        direction = None
        reUnit = u'[^\W\d]+'
        for arg in args:
            if arg is None:
                pass
            elif type(arg) is int:
                direction = arg
            elif type(arg) is str or type(arg) is unicode:
                reUnit = arg
            else:
                raise TypeError('unexpected argument type')
        reUnit = '(' + reUnit + ')'
        # digitRegex = r'([\d.,]+)'
        digitRegex = r'([+-]?[.,]?\d+[\d.,]*)'
        if direction is None:
            forward = self._reValueUnit(reUnit, Direction.forward)
            backward = self._reValueUnit(reUnit, Direction.backward)
            return forward + backward
        elif direction == Direction.forward:
            regex = digitRegex + r'\s*' + reUnit
        elif direction == Direction.backward:
            regex = reUnit + r'\s*' + digitRegex
        else:
            raise DirectionError('Direction should by Direction.forward or' \
                                 + ' Direction.backward')
        result = self._regex(regex)
        if direction == Direction.backward:
            resultLen = len(result)
            i = 0
            while i < resultLen:
                (unit, value) = result[i]
                result[i] = (value, unit)
                i += 1
        return result


def removeWrapAndClone(elem):
    if type(elem) is list:
        elem = elem[0]
        elem = elem._clone()
        elem._structure = Structure.list
    elif type(elem) is dict:
        key = list(elem.keys())
        key = key[0]
        elem = elem[key]
        elem = elem._clone()
        elem._structure = Structure.dict
        elem._key = key
    else:
        elem = elem._clone()
        elem._structure = Structure.single
    return elem


def getHash(instance):
    if isinstance(instance, ParsBase):
        return instance._getHash()
    elif type(instance) is list:
        hashList = []
        for elem in instance:
            hashList.append(elem._getHash())
        hashStr = ''
        for h in sorted(hashList):
            hashStr += h
        return md5(hashStr)
    elif type(instance) is dict:
        hashList = []
        for key in instance:
            hashList.append(instance[key]._getHash())
        hashStr = ''
        for h in sorted(hashList):
            hashStr += h
        return md5(hashStr)
    else:
        raise ParsError('instance must be ParsBase or list ParsBase' + \
                        + ' or dict ParsBase')


class ParsBase(object):

    _encoding = 'utf-8'
    _NoneObjectUnicode = u'None'
    _saveInstanceDefault = True
    _maxAttemptsDefault = 3

    @staticmethod
    def _unicodePostProcessingDefault(unistr):
        return unistr

    @staticmethod
    def _breakerDefault():
        pass

    _unicodePostProcessing = _unicodePostProcessingDefault

    def __init__(self, query, replaceObj=None, **kwargs):
        self._query = query
        self.parent = None
        self._childNames = set()
        self._structure = Structure.single
        self._name = None
        self._url_local = ''
        self._url_scheme_local = ''
        self._url_netloc_local = ''
        self._url_path_local = ''
        self._url_query_local = ''
        self._url_fragment_local = ''
        self._queryArgs = ()
        self._queryKwargs = {}
        self._key = None
        self._elem = None
        if replaceObj is not None:
            replaceObj = replaceObj._clone()
        self._replaceObj = replaceObj
        self._catcher = kwargs.pop('catcher', None)
        self._needControl = kwargs.pop('needControl', False)
        maxAttempts = kwargs.pop('maxAttempts', None)
        if maxAttempts is not None:
            self._maxAttempts = maxAttempts
        self._breaker = kwargs.pop('breaker', ParsBase._breakerDefault)
        self._maxIteration = kwargs.pop('maxIteration', None)
        self._hash = None
        self._treeHash = None
        self._catcherCalled = False
        self._treeKey = kwargs.pop('treeKey', True)

    def _replaceAttributes(self, fromInstance):
        self.parent = fromInstance.parent
        self._structure = fromInstance._structure
        self._name = fromInstance._name
        if self._url_scheme_local == '':
            self._url_scheme_local = fromInstance._url_scheme_local
        if self._url_netloc_local == '':
            self._url_netloc_local = fromInstance._url_netloc_local
        if self._url_path_local == '':
            self._url_path_local = fromInstance._url_path_local
        if self._url_query_local == '':
            self._url_query_local = fromInstance._url_query_local
        if self._url_fragment_local == '':
            self._url_fragment_local = fromInstance._url_fragment_local

    def __getattribute__(self, name):
        if name == '_unicode':
            result = object.__getattribute__(self, name)
            return self._unicodePostProcessing(result)
        else:
            return object.__getattribute__(self, name)

    def __getattr__(self, name):
        if name == '_unicodePostProcessing':
            return ParsBase._unicodePostProcessing
        elif name == '_saveInstance':
            return ParsBase._saveInstanceDefault
        elif name == '_maxAttempts':
            return ParsBase._maxAttemptsDefault
        else:
            raise AttributeError(name)

    def __setattr__(self, name, value):
        if name[0] != '_' and name != 'parent': # and isParsStructure(value):
            if not self._childNameExists(name):
                return self._setChild(value, name)
            return object.__setattr__(self, name, value)
        if name in ('_url_local', '_url_scheme_local', '_url_netloc_local'
                    , '_url_path_local', '_url_query_local'
                    , '_url_fragment_local') and value is None:
            value = ''
        if name == '_url_local':
            self._url_scheme, self._url_netloc, self._url_path, self._url_query\
                , self._url_fragment = urlparse.urlsplit(value)
        if name == '_structure':
            if value not in (Structure.single, Structure.list, Structure.dict):
                raise StructureError('Structure should by Structure.single or \
                                    Structure.list')
        object.__setattr__(self, name, value)

    def __unicode__(self):
        if getattr(self, '_elem', None) is None:
            return ParsBase._NoneObjectUnicode
        return self._unicode

    def __str__(self):
        return self.__unicode__().encode(ParsBase._encoding)

    def __repr__(self):
        return self.__str__()

    @property
    def _url(self):
        return urlparse.urlunsplit((self._url_scheme, self._url_netloc,
                                   self._url_path, self._url_query,
                                   self._url_fragment))

    @_url.setter
    def _url(self, value):
        self._url_local = value

    @property
    def _url_scheme(self):
        result = self._url_scheme_local
        if result == '':
            if self.parent is not None:
                result = self.parent._url_scheme
        return result

    @_url_scheme.setter
    def _url_scheme(self, value):
        self._url_scheme_local = value

    @property
    def _url_netloc(self):
        result = self._url_netloc_local
        if result == '':
            if self._url_scheme_local == '':
                if self.parent is not None:
                    result = self.parent._url_netloc
        return result

    @_url_netloc.setter
    def _url_netloc(self, value):
        self._url_netloc_local = value

    @property
    def _url_path(self):
        result = self._url_path_local
        if result == '':
            if self._url_scheme_local == self._url_netloc_local == '':
                if self.parent is not None:
                    result = self.parent._url_path
        return result

    @_url_path.setter
    def _url_path(self, value):
        self._url_path_local = value

    @property
    def _url_query(self):
        result = self._url_query_local
        if result == '':
            if self._url_scheme_local == self._url_netloc_local ==\
                    self._url_path_local == '':
                if self.parent is not None:
                    result = self.parent._url_query
        return result

    @_url_query.setter
    def _url_query(self, value):
        self._url_query_local = value

    @property
    def _url_fragment(self):
        result = self._url_fragment_local
        if result == '':
            if self._url_scheme_local == self._url_netloc_local ==\
                    self._url_path_local == '':
                if self.parent is not None:
                    result = self.parent._url_fragment
        return result

    @_url_fragment.setter
    def _url_fragment(self, value):
        self._url_fragment_local = value

    def _selfUrl(self):
        return [self._url]

    def _clone(self):
        clone = copy.copy(self)
        clone._childNames = copy.copy(self._childNames)
        if clone._replaceObj is not None:
            clone._replaceObj = clone._replaceObj._clone()
        return clone

    def _childNameExists(self, name):
        if self._replaceObj is not None:
            return self._replaceObj._childNameExists(name)
        return hasattr(self, name)

    def _setChild(self, child, name=None):
        if self._replaceObj is not None:
            return self._replaceObj._setChild(child, name)
        child = removeWrapAndClone(child)
        if name is None:
            name = child._name
        else:
            child._name = name
        if name is None:
            # print(self._name)
            raise UndefinedChildName()
        if name in self._childNames:
            raise DuplicateChildName(name)
        if hasattr(self, name):
            raise SetChildError('Attribute ' + name + ' already exists')
        result = object.__setattr__(self, name, child)
        self._childNames.add(name)
        return result

    def __call__(self, *args, **kwargs):
        clone = self._clone()
        childs = []
        queryArgs = []
        for arg in args:
            if isParsStructure(arg):
                childs.append(arg)
            else:
                queryArgs.append(arg)
        clone._queryArgs = tuple(queryArgs)
        clone._queryKwargs = kwargs
        for child in childs:
            clone._setChild(child)
        return clone

    def _runQuery(self, query=None):
        if query is None:
            query = self._query
        if query is None:
            raise UndefinedQuery
        return query(self, *self._queryArgs, **self._queryKwargs)

    def _processing(self):
        pass

    def _calcHash(self):
        return md5(str(self))

    def _getHash(self):
        if self._hash is None:
            selfHash = self._calcHash()
            if type(selfHash) is unicode:
                selfHash = selfHash.encode(ParsBase._encoding)
            elif type(selfHash) is not str:
                selfHash = str(selfHash)
            self._hash = selfHash
        return self._hash

    def _calcListTreeHash(self, childList):
        hashList = []
        for child in childList:
            hashList.append(child._getTreeHash())
        hashStr = ''
        for h in sorted(hashList):
            hashStr += h
        return md5(hashStr)

    def _calcDictTreeHash(self, childDict):
        hashList = []
        for key in childDict:
            hashList.append(childDict[key]._getTreeHash())
        hashStr = ''
        for h in sorted(hashList):
            hashStr += h
        return md5(hashStr)

    def _calcTreeHash(self):
        hashList = []
        for childName in self._childNames:
            child = self[childName]
            if type(child) is list:
                hashList.append(self._calcListTreeHash(child))
            elif type(child) is dict:
                hashList.append(self._calcDictTreeHash(child))
            else:
                hashList.append(child._getTreeHash())
        hashStr = ''
        for h in sorted(hashList):
            hashStr += h
        hashStr = self._getHash() + hashStr
        return md5(hashStr)

    def _getTreeHash(self):
        if not self._treeKey:
            return ''
        if self._treeHash is None:
            self._treeHash = self._calcTreeHash()
        return self._treeHash

    def _goodChilds(self):
        goodChilds = True
        for childName in self:
            child = self[childName]
            if child is None:
                goodChilds = False
            if isinstance(child, NoneObject):
                goodChilds = False
            if type(child) is list:
                if len(child) == 0:
                    goodChilds = False
            if type(child) is dict:
                if len(child) == 0:
                    goodChilds = False
            if isinstance(child, ParsBase) \
                    and not isinstance(child, NoneObject):
                if child._elem is None:
                    goodChilds = False
        return goodChilds

    def __eq__(self, instance):
        if self._getHash() == instance._getHash():
            return True
        else:
            return False

    def __ne__(self, instance):
        return not self.__eq__(instance)

    def _compareWithChild(self, instance):
        if self != instance:
            return False
        if self._childNames != instance._childNames:
            return False
        for childName in self:
            if self[childName] != instance[childName]:
                return False
        return True

    def _findIdenticalInstances(self, instances):
        resultList = []
        if instances is None:
            return resultList
        for instance in instances:
            if self._compareWithChild(instance):
                resultList.append(instance)
        return resultList

    def _instanceControl(self, oldAttemptsInstances=None):
        if oldAttemptsInstances is None or oldAttemptsInstances == []:
            if self._goodChilds() and self._elem is not None \
                    and len(self._childNames) > 0:
                return ControlResult.ok
            else:
                return ControlResult.bad
        else:
            if len(oldAttemptsInstances) >= self._maxAttempts:
                return ControlResult.fault
            if self._goodChilds() and self._elem is not None \
                    and len(self._childNames) > 0:
                return ControlResult.ok
            for oldInstance in oldAttemptsInstances:
                if self._compareWithChild(oldInstance):
                    return ControlResult.ok
            return ControlResult.bad

    def _noneConstruct(self):
        if self._structure == Structure.list:
            return []
        elif self._structure == Structure.dict:
            return {}
        elif self._structure == Structure.single:
            instance = self._clone()
            instance._elem = None
            for childName in instance:
                instance[childName].parent = instance
                instance[childName] = instance[childName]._noneConstruct()
            return instance
        else:
            raise StructureError('Structure should by Structure.single or \
                                 Structure.list or Structure.dict')

    def _instanceConstruct(self, queryResult
                           , saveInstance=False
                           , oldAttemptsInstances=None):
        saveInstanceOriginal = saveInstance
        saveInstance = saveInstance or self._saveInstance
        catcher = self._catcher
        if catcher is None:
            saveChild = False
        else:
            saveChild = True
        saveChild = saveChild or saveInstance
        instance = self._clone()
        instance._elem = queryResult
        if oldAttemptsInstances is None or oldAttemptsInstances == []:
            instance._processing()
        else:
            instance._processing(oldAttemptsInstances)
        saveChild = saveChild or instance._needControl
        if instance._structure == Structure.dict:
            key = instance._key
            if key is not None:
                key.parent = instance
                key = key._construct(saveInstance=True)
                key = key._elem
        if instance._replaceObj is not None:
            instance._replaceObj.parent = instance
            newInstance = instance._replaceObj._construct(saveInstance=True)
            newInstance._replaceAttributes(instance)
            instance = newInstance
        for childName in instance:
            instance[childName].parent = instance
            if instance._elem is None:
                if saveChild and (self._structure == Structure.single \
                                  or (self._structure == Structure.dict \
                                      and key is not None)):
                    instance[childName] = instance[childName]._noneConstruct()
            else:
                instance[childName] = instance[childName]._construct(saveChild)
        if instance._needControl:
            controlResult = instance._instanceControl(oldAttemptsInstances)
            if controlResult  == ControlResult.fault:
                instance = NoneObject(self)
            elif controlResult == ControlResult.bad:
                if oldAttemptsInstances is None:
                    oldAttemptsInstances = []
                oldAttemptsInstances.append(instance)
                instance = self._instanceConstruct(
                    queryResult = queryResult
                    , saveInstance = saveInstanceOriginal
                    , oldAttemptsInstances = oldAttemptsInstances
                )
            elif controlResult == ControlResult.ok:
                self._breaker()
            else:
                raise ControlResultError('ControlResult should by' \
                                         + ' ControlResult.ok or' \
                                         + ' ControlResult.bad or' \
                                         + ' ControlResult.fault')
        else:
            self._breaker()
        if (catcher is not None) and (not isinstance(instance, NoneObject)) \
                and instance is not None and instance._elem is not None:
            if not instance._catcherCalled:
                try:
                    catcher(instance)
                except DuplicateTree:
                    pass
                instance._catcherCalled = True
        if self._structure == Structure.dict:
            if key is None and (instance is None or instance._elem is None):
                return None, None
            if saveInstance:
                return key, instance
            else:
                return None, None
        elif self._structure == Structure.list \
                and (instance is None or instance._elem is None):
            return None
        else:
            if saveInstance:
                return instance
            else:
                return None

    def _listInstanceConstruct(self, listQueryResult, saveInstance=False):
        saveInstance = saveInstance or self._saveInstance
        instanceList = []
        iterNumber = 0
        for queryResult in listQueryResult:
            if self._maxIteration is not None:
                iterNumber += 1
                if iterNumber > self._maxIteration:
                    break
            instance = self._instanceConstruct(queryResult, saveInstance)
            if saveInstance and instance is not None \
                    and instance._elem is not None:
                instanceList.append(instance)
        if saveInstance:
            return instanceList
        else:
            return None

    def _dictInstanceConstruct(self, queryResult, saveInstance=False):
        saveInstance = saveInstance or self._saveInstance
        instanceDict = {}
        resultLen = len(queryResult)
        i = 0
        iterNumber = 0
        while i < resultLen:
            if self._maxIteration is not None:
                iterNumber += 1
                if iterNumber > self._maxIteration:
                    break
            key, instance = self._instanceConstruct(queryResult[i]
                                                    , saveInstance)
            if key is None and instance is None:
                i += 1
                continue
            if key is None:
                key = i
            if saveInstance and instance is not None:
                instanceDict[key] = instance
            i += 1
        if saveInstance:
            return instanceDict
        else:
            return None

    def _prepareQueryResult(self, queryResult):
        return queryResult

    def _construct(self, saveInstance=False):
        saveInstance = saveInstance or self._saveInstance
        queryResult = self._runQuery()
        queryResult = self._prepareQueryResult(queryResult)
        if type(queryResult) is not list:
            raise BadQueryResult('The query result should be a list')
        if self._structure == Structure.single:
            if len(queryResult) == 0:
                # result = NoneObject(self)
                result = self._noneConstruct()
            elif len(queryResult) == 1:
                result = self._instanceConstruct(queryResult[0], saveInstance)
            else:
                # msg = 'A list of query results must contain one element'
                # raise BadQueryResult(msg, queryResult=queryResult
                                     # , errorObj=self)
                result = self._noneConstruct()
        elif self._structure == Structure.list:
            result = self._listInstanceConstruct(queryResult, saveInstance)
        elif self._structure == Structure.dict:
            result = self._dictInstanceConstruct(queryResult, saveInstance)
        else:
            raise StructureError('Structure should by Structure.single or \
                                 Structure.list or Structure.dict')
        if saveInstance:
            return result
        else:
            return None

    def __iter__(self):
        return iter(self._childNames)

    def __getitem__(self, index):
        return getattr(self, index)

    def __setitem__(self, index, value):
        setattr(self, index, value)

    def __contains__(self, x):
        return x in self._childNames

    def print(self):
        print(self._name, '=', self)

    def printTree(self, fileName=None, fileObject=sys.stdout, level=0, index=None):
        if fileName is not None and fileObject == sys.stdout:
            outStream = open(fileName, 'a')
        else:
            outStream = fileObject
        tabSpace = '    '
        indent = tabSpace * level
        # indentListSymbol = indent + ' ' * (len(tabSpace) - 2)
        indentListSymbol = indent + ' ' * (len(tabSpace) - 0)
        selfStr = self.__str__()
        selfStr = selfStr.replace('\n', '')
        selfName = self._name
        if index is not None:
            if type(index) is unicode:
                index = index.encode(ParsBase._encoding)
            else:
                index = str(index)
            selfName += '[' + index + ']'
        print(indent + selfName + ': ' + selfStr, file=outStream)
        for childName in sorted(self._childNames):
            child = self[childName]
            if type(child) is list:
                print(indentListSymbol + childName + ': [', file=outStream)
                i = 0
                listLen = len(child)
                while i < listLen:
                    child[i].printTree(fileName=fileName, fileObject=outStream
                                       , level=level+2, index=i)
                    i += 1
                print(indentListSymbol + ']', file=outStream)
            elif type(child) is dict:
                print(indentListSymbol + childName + ': {', file=outStream)
                for key in sorted(child):
                    child[key].printTree(fileName=fileName, fileObject=outStream
                                         , level=level+2, index=key)
                print(indentListSymbol + '}', file=outStream)
            else:
                child.printTree(fileName=fileName, fileObject=outStream
                                , level=level+1)
        if fileName is not None and fileObject == sys.stdout:
            outStream.close()

    def printPath(self):
        if self.parent is not None:
            self.parent.printPath()
        string = self._name + ': ' + str(self)
        print(string)

    @property
    def _unicode(self):
        return unicode(self._elem)


class NoneObject(ParsBase):
    def __init__(self, fromInstance=None):
        query = None
        ParsBase.__init__(self, query)
        if fromInstance is not None:
            self._replaceAttributes(fromInstance)
        self._elem = None

    @property
    def _unicode(self):
        return ParsBase._NoneObjectUnicode


class Container:
    pass


class Processor(object):

    def __init__(self):
        self._result = Container()

    @property
    def result(self):
        return self._result

    def __setattr__(self, name, value):
        if issubclass(value.__class__, ParsBase):
            value._name = name
        object.__setattr__(self, name, value)

    def __call__(self, queryRoot):
        queryRoot = removeWrapAndClone(queryRoot)
        resultName = queryRoot._name
        if resultName is None:
            raise UndefinedQueryRootName()
        result = queryRoot._construct()
        setattr(self._result, resultName, result)


class ElemTypeMixin(object):

    def printElemType(self):
        print(type(self._elem))


class ListElemTypeMixin(object):

    def printElemType(self):
        print('[', end='')
        sep = ''
        for elem in self._elem:
            print(sep, end='')
            print(type(elem), end='')
            sep = ', '
        print(']')


class UnicodeListMixin(object):

    @property
    def _unicode(self):
        unistr = u'[\n'
        for elem in self._elem:
            unistr += etree.tounicode(elem, method='xml', pretty_print=True
                                      , with_tail=False) + u'\n,\n'
        unistr = unistr[0:-2] + u']'
        return unistr


class UnicodeTextMixin(object):

    @property
    def _unicode(self):
        return etree.tounicode(self._elem, method='text', pretty_print=False
                               , with_tail=False)


class UnicodeStrMixin(object):

    @property
    def _unicode(self):
        return self._elem


class UnicodeTreeMixin(object):

    @property
    def _unicode(self):
        return etree.tounicode(self._elem, method='html', pretty_print=True
                               , with_tail=True)


class TreeXpath(UnicodeTreeMixin, XpathQueryMixin, RegexQueryMixin, ParsBase):
    pass


class Str(UnicodeStrMixin, RegexQueryMixin, ParsBase):

    def _processing(self):
        elem = self._elem
        if type(elem) is unicode:
            pass
        if type(elem) is str:
            elem = unicode(elem.decode(ParsBase._encoding))
        else:
            self._elem = unicode(elem)

    @property
    def _unicode(self):
        if self._elem is None:
            return u'None'
        return self._elem

    def __int__(self):
        return int(self._elem)


class Value(Str):

    def _processing(self):
        self._elem = self._elem[0]
        Str._processing(self)


class Unit(Str):

    def _processing(self):
        self._elem = self._elem[1]
        Str._processing(self)


class Tail(Str):

    def _processing(self):
        self._elem = unicode(self._elem.tail)


class Text(Str):

    def _processing(self):
        self._elem = unicode(self._elem.text)


class Url(Str):

    def _processing(self):
        Str._processing(self)
        self._url = self._elem
        self._elem = self._url


class Int(ParsBase):

    def _processing(self):
        # intFunc = self._int
        # print(intFunc.func_name)
        # print(dir(intFunc))
        # print(type(intFunc))
        self._elem = self._int(self._elem)

    @staticmethod
    def _intDefault(number):
        return int(number)

    _int = _intDefault

    def __int__(self):
        return self._elem

    # def __getattr__(self, name):
        # if name == '_int':
            # return Int._int
        # return ParsBase.__getattr__(self, name)


class HtmlElements(UnicodeStrMixin, ParsBase):

    def _prepareQueryResult(self, queryResult):
        return [queryResult]

    def _processing(self):
        str = u''
        for elem in self._elem:
            str += etree.tounicode(elem, pretty_print=False,
                                   method='html')
        self._elem = str


class WebPage(object):

    __module__ = os.path.splitext(os.path.basename(__file__))[0]

    pageConfirmedDefault = True

    def __init__(self, page=None, **kwargs):
        self.page = page
        pageConfirmed = kwargs.get('pageConfirmed', None)
        if pageConfirmed is not None:
            self.pageConfirmed = pageConfirmed
        self.proxy = kwargs.get('proxy', None)
        self.url = kwargs.get('url', None)
        self.uuid = uuid.uuid4()
        self.proxyEventRegistered = False

    def __getattr__(self, name):
        if name == 'pageConfirmed':
            return WebPage.pageConfirmedDefault
        if name == 'proxy':
            return None
        else:
            raise AttributeError(name)

    def regProxyGoodPage(self):
        if self.proxyEventRegistered:
            return
        if self.proxy is not None:
            self.proxy.regSuccessRequest(self.url)
            self.proxyEventRegistered = True

    def regProxyBadPage(self):
        if self.proxyEventRegistered:
            return
        if self.proxy is not None:
            self.proxy.regFailedRequest(self.url, BadPageError())
            self.proxyEventRegistered = True

    @property
    def httpCode(self):
        try:
            return self.page.response.code
        except AttributeError:
            return None

    @property
    def httpBody(self):
        try:
            return self.page.response.body
        except AttributeError:
            return None

    @property
    def response(self):
        try:
            return self.page.response
        except AttributeError:
            return None


class Proxy(object):

    __module__ = os.path.splitext(os.path.basename(__file__))[0]

    def __init__(self, address, port, type, user=None, password=None):
        self.address = address
        self.port = port
        self.type = type
        self.user = user
        self.password = password
        self.requests = 0
        self.successRequests = 0
        self.failedRequests = 0
        self.failed = False

    def printInfo(self, event):
        def attrStr(self, attr):
            return attr + ': ' + str(getattr(self, attr, None))
        res = 'EVENT: ' + event + '\n'
        res += attrStr(self, 'address') + '\n'
        res += attrStr(self, 'port') + '\n'
        res += attrStr(self, 'type') + '\n'
        res += attrStr(self, 'requests') + '\n'
        res += attrStr(self, 'successRequests') + '\n'
        res += attrStr(self, 'failedRequests') + '\n'
        res += attrStr(self, 'failed') + '\n'
        res += 'proxyCount: ' + str(Web._proxyCount()) + '\n'
        if self.requests == 0:
            res += 'proxyRatio: ' + 'undefined' + '\n'
        else:
            proxyRatio = float(self.successRequests)/float(self.requests)
            res += 'proxyRatio: ' + repr(proxyRatio) + '\n'
        res += '---------------------------------------------\n'
        f = open('var/log/proxyInfo.info', 'a')
        f.write(res)
        f.close()

    def regSuccessRequest(self, url):
        # self.printInfo('successStart')
        self.requests += 1
        self.successRequests += 1
        proxy = self.address + ':' + self.port
        event = 'successRequest'
        Web._writeLogProxyEvent(proxy, event, url)
        # self.printInfo('successEnd')

    def regFailedRequest(self, url, exception):
        # self.printInfo('failedStart')
        self.requests += 1
        self.failedRequests += 1
        if not self.failed and self.failedRequests > Web.maxFailedProxyRequests:
            if self.successRequests == 0:
                self.failed = True
                Web._setProxyFailed(self)
            elif Web._proxyCount() > Web.minimumProxies:
                proxyRatio = float(self.successRequests)/float(self.requests)
                if proxyRatio < Web.proxyRejectRatio:
                    self.failed = True
                    Web._setProxyFailed(self)
        if isinstance(exception, ProxyServerError):
            event = 'proxyServerError'
        elif isinstance(exception, BadPageError):
            event = 'badPage'
        elif isinstance(exception, grab.error.GrabTimeoutError):
            event = 'timeoutError'
        elif isinstance(exception, WebClientError) \
                or isinstance(exception, WebServerError):
            httpCode = exception.httpCode
            if httpCode is None:
                httpCode = 'Error'
            else:
                httpCode = str(httpCode)
            event = 'http' + httpCode
        elif isinstance(exception, PageControlFault):
            event = 'pageControlFault'
        elif isinstance(exception, grab.error.GrabConnectionError):
            event = 'connectionError'
        elif isinstance(exception, grab.error.GrabNetworkError):
            event = 'networkError'
        else:
            event = 'failedRequest'
        proxy = self.address + ':' + self.port
        Web._writeLogProxyEvent(proxy, event, url)
        # self.printInfo('failedEnd')


class HttpCodeCheck(object):
    # TODO optimize class based statistical proxies

    def __init__(self):
        self.codeStat = {}
        self.allCodeEvents = 0

    def regCodeEvent(self, code):
        if code is None:
            return
        codeCnt = self.codeStat.get(code, 0)
        codeCnt += 1
        self.codeStat[code] = codeCnt
        self.allCodeEvents += 1

    def __call__(self, code):
        if code is None:
            return None
        result = None
        self.regCodeEvent(code)
        codeEvens = self.codeStat[code]
        allCodeEvents = self.allCodeEvents
        if code == 404:
            if codeEvens >= 4:
                codeWeight = float(codeEvens)/float(allCodeEvents)
                if codeWeight >= 0.74:
                    return code
        return result


class Web(object):

    _proxies = None
    _failedProxies = set()
    proxyFile = None
    hammerTimeouts = None
    randomDelayPeriod = None
    considerTimeoutProxyError = True
    maxFailedProxyRequests = 5
    logDir = None
    proxyStatDir = None
    _proxyStatFileName = None
    _lastReqTime = 0
    attempts = 5
    errorLogDir = None
    _errorLogFilename = None
    serverErrorWait = 120
    minimumProxies = 10
    proxyRejectRatio = 0.5
    allow404 = False
    log404dir = None
    _log404fileName = None

    @staticmethod
    def _nonePageControlFunc(page, url, proxy):
        pass

    pageControlFunc = _nonePageControlFunc

    @staticmethod
    def getGrabPage(url, oldAttemptsPages=None):
        exceptProxy = set()
        if oldAttemptsPages is not None and len(oldAttemptsPages) > 0:
            for oldPage in oldAttemptsPages:
                if oldPage.proxy is not None:
                    exceptProxy.add(oldPage.proxy)
        if Web._proxyMode():
            if Web._proxies is None:
                Web._loadProxies()
            page = Web._getGrabPageProxy(url, exceptProxy)
        else:
            page = Web._getGrabPageDirect(url)
        if page.httpCode == 404:
            Web._writeLog404(url)
        return page

    @staticmethod
    def _getGrabPageProxy(url, exceptProxy=None, waitCnt=0, httpCodeCheck=None):
        if httpCodeCheck is None:
            httpCodeCheck = HttpCodeCheck()
        if exceptProxy is None:
            exceptProxy = set()
        proxy = Web._nextProxy(exceptProxy)
        kwargs = {}
        kwargs['proxy'] = proxy.address + ':' + proxy.port
        kwargs['proxy_type'] = proxy.type
        if proxy.user is not None:
            password = proxy.password
            if password is None:
                password = ''
            kwargs['proxy_userpwd'] = proxy.user + ':' + password
        try:
            page = Web._loadGrabPage(url, proxy, **kwargs)
            if page.pageConfirmed:
                # proxy.regSuccessRequest(url)
                page.regProxyGoodPage()
        except (WebError, grab.error.GrabNetworkError) as e:
            if isinstance(e, WebError):
                codeCheck = httpCodeCheck(e.httpCode)
                page = e.page
                if codeCheck is not None and page is not None:
                    if (codeCheck == 404 and Web.allow404):
                        if page.pageConfirmed:
                            # proxy.regSuccessRequest(url)
                            page.regProxyGoodPage()
                        return page
                if e.httpCode in (503, 507, 509):
                    if waitCnt > 0:
                        time.sleep(Web.serverErrorWait)
                    waitCnt += 1
            exceptProxy.add(proxy)
            try:
                page = Web._getGrabPageProxy(url, exceptProxy
                                             , waitCnt, httpCodeCheck)
                regFailed = True
                if isinstance(e, WebError):
                    if e.httpCode == page.httpCode == 404 and Web.allow404:
                        regFailed = False
                if regFailed:
                    proxy.regFailedRequest(url, e)
                else:
                    page.regProxyGoodPage()
            except AllProxyAlreadyUsed as e1:
                exceptionList = getattr(e1, 'exceptionList', None)
                if exceptionList is None:
                    e1.exceptionList = []
                e1.exceptionList.append(e)
                raise e1
        return page

    @staticmethod
    def _getGrabPageDirect(url):
        attempts = Web.attempts
        if attempts is None:
            attempts = 1
        i = 0
        while i < attempts:
            try:
                return Web._loadGrabPage(url)
            except grab.error.GrabNetworkError as e:
                i += 1
                if i < attempts:
                    time.sleep(Web.serverErrorWait)
                    continue
                raise
            except WebServerError as e:
                if e.httpCode in (503, 507, 509):
                    i += 1
                    if i < attempts:
                        time.sleep(Web.serverErrorWait)
                        continue
                raise
            except WebClientError as e:
                if e.httpCode == 404 and Web.allow404:
                    return e.page
                raise

    @staticmethod
    def _nextProxy(exceptProxy=None):
        if exceptProxy is None:
            exceptProxy = set()
        if Web._proxies is None:
            raise ProxyError('set of proxy servers is not initialized')
        if len(Web._proxies) == 0:
            if len(Web._failedProxies) == 0:
                raise ProxyError('set of proxy servers is empty')
            else:
                raise ProxyError('All proxy servers is failed')
        proxies = Web._proxies - exceptProxy
        if len(proxies) == 0:
            raise AllProxyAlreadyUsed()
        return reduce((lambda x, y: x if x.requests <= y.requests else y)
                      , proxies)

    @staticmethod
    def _setProxyFailed(proxy):
        Web._failedProxies.add(proxy)
        Web._proxies.remove(proxy)

    @staticmethod
    def _writeLogError(url, proxy, exception):
        if Web.errorLogDir is None:
            return
        fileName = Web._errorLogFilename
        if fileName is None:
            dateName = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            fileName = dateName + '.' + str(os.getpid()) + '.neterror.log'
            mkdirs(Web.errorLogDir)
            fileName = Web.errorLogDir + fileName
            Web._errorLogFilename = fileName
        logFile = open(fileName, 'a')
        if proxy is not None:
            proxy = proxy.address + ':' + proxy.port
        else:
            proxy = 'None'
        proxy = '<<PROXY>> ' + proxy
        error = '<<ERROR>> '
        error += className(exception) + ': '
        error += str(exception)
        if type(url) is unicode:
            url = url.encode(ParsBase._encoding)
        url = '<<URL>> ' + '"' + url + '"'
        string = proxy + '\t' + error + '\t' + url
        string = delElements(string, '\r\n')
        string += '\n'
        if type(string) is unicode:
            string = string.encode(ParsBase._encoding)
        logFile.write(string)
        logFile.close()

    @staticmethod
    def _writeLogProxyEvent(proxy, event, url):
        if Web.proxyStatDir is None:
            return
        fileName = Web._proxyStatFileName
        if fileName is None and Web.proxyStatDir is not None:
            dateName = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            fileName = dateName + '.' + str(os.getpid()) + '.proxystat'
            mkdirs(Web.proxyStatDir)
            fileName = Web.proxyStatDir + fileName
            Web._proxyStatFileName = fileName
        statFile = open(fileName, 'a')
        string = '<' + event + '>\t' + proxy + '\t' + '"' + url + '"' + '\n'
        if type(string) is unicode:
            string = string.encode(ParsBase._encoding)
        statFile.write(string)
        statFile.close()

    @staticmethod
    def _writeLog404(url):
        if Web.log404dir is None:
            return
        fileName = Web._log404fileName
        if fileName is None:
            dateName = datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            fileName = dateName + '.' + str(os.getpid()) + '.404.log'
            mkdirs(Web.log404dir)
            fileName = Web.log404dir + fileName
            Web._log404fileName = fileName
        logFile = open(fileName, 'a')
        if type(url) is unicode:
            url = url.encode(ParsBase._encoding)
        logFile.write(url+'\n')
        logFile.close()

    @staticmethod
    def _proxyMode():
        return (Web._proxies is not None) or (Web.proxyFile is not None)

    @staticmethod
    def _loadProxies():
        proxyFile = open(Web.proxyFile, 'r')
        wrongTypeFile = None
        proxies = set()
        typePattern = RegexCache.compile(r'^(\w+)\t[\w:@.]+\s*$')
        addrPortPattern = RegexCache.compile(r'[\t@]([\w.]+):(\d+)\s*$')
        usrPwdRegex = r'^\w+\t(\w+):([^\s:]+)*@[\w.]+:\d+\s*$'
        usrPwdPattern = RegexCache.compile(usrPwdRegex)
        for line in proxyFile:
            proxyType = typePattern.findall(line)[0]
            proxyType = proxyType.lower()
            if proxyType not in ('http', 'socks4', 'socks5'):
                if Web.logDir is not None and wrongTypeFile is None:
                    mkdirs(Web.logDir)
                    wrongTypeFile = open(Web.logDir+'wrongProxyType.log', 'a')
                if wrongTypeFile is not None:
                    wrongTypeFile.write(Web.proxyFile + ' : ' + line)
                continue
            proxyAddress, proxyPort = addrPortPattern.findall(line)[0]
            usrPwd = usrPwdPattern.findall(line)
            if len(usrPwd) == 1:
                proxyUser, proxyPassword = usrPwd[0]
            else:
                proxyUser = proxyPassword = None
            proxy = Proxy(proxyAddress, proxyPort, proxyType
                          , proxyUser, proxyPassword)
            proxies.add(proxy)
        Web._proxies = proxies
        if wrongTypeFile is not None:
            wrongTypeFile.close()
        proxyFile.close()

    @staticmethod
    def _lastRequestTime(setTime=None):
        if setTime is None:
            return Web._lastReqTime
        else:
            Web._lastReqTime = setTime
            return setTime

    @staticmethod
    def _proxyCount():
        return len(Web._proxies)

    @staticmethod
    def _grabPageHttpCodeCheck(page, proxy):
        response = getattr(page, 'response', None)
        code = None
        body = None
        if response is not None:
            code = getattr(response, 'code', None)
            body = getattr(response, 'body', None)
        if 100 <= code < 200:
            raise WebClientError(httpCode=code, httpBody=body, page=page)
        elif 200 <= code < 300:
            if code == 206:
                raise WebClientError(httpCode=code, httpBody=body, page=page)
            return
        elif 300 <= code < 400:
            raise WebClientError(httpCode=code, httpBody=body, page=page)
        elif 400 <= code < 500:
            raise WebClientError(httpCode=code, httpBody=body, page=page)
        elif 500 <= code < 600:
            if code == 504 and proxy is not None:
                raise ProxyServerError(httpCode=code, httpBody=body, page=page)
            raise WebServerError(httpCode=code, httpBody=body, page=page)
        else:
            raise WebClientError(httpCode=code, httpBody=body, page=page)

    @staticmethod
    def _loadGrabPage(url, currentProxy=None, **kwargs):
        # _breaker(5)
        kwargs = kwargs.copy()
        page = Grab()
        if Web.hammerTimeouts is not None:
            kwargs['hammer_mode'] = True
            kwargs['hammer_timeouts'] = Web.hammerTimeouts
        if len(kwargs) > 0:
            page.setup(**kwargs)
        if Web.randomDelayPeriod is not None:
            delaySecFrom, delaySecTo = Web.randomDelayPeriod
            sleepSeconds = random.uniform(delaySecFrom, delaySecTo)
            timeFromLastRequest = time.time() - Web._lastRequestTime()
            sleepSeconds -= timeFromLastRequest
            if sleepSeconds > 0:
                time.sleep(sleepSeconds)
        try:
            try:
                requestTime = time.time()
                page.go(url)
                page = WebPage(page, proxy=currentProxy, url=url)
                Web._grabPageHttpCodeCheck(page, currentProxy)
                Web.pageControlFunc(page, url, currentProxy)
                Web._lastRequestTime(requestTime)
            except grab.error.GrabConnectionError as e:
                if currentProxy is None:
                    raise
                strerror = getattr(e, 'strerror', None)
                if strerror is not None:
                    if strerror.find(currentProxy.address) != -1:
                        raise ProxyServerError(originalException=e)
                    if strerror.startswith(r"Can't complete SOCKS4 connection to"):
                        raise ProxyServerError(originalException=e)
                    if strerror.startswith(r"Failed to receive SOCKS4"):
                        raise ProxyServerError(originalException=e)
                raise
        except (grab.error.GrabNetworkError, WebError) as e:
            Web._writeLogError(url, currentProxy, e)
            raise
        return page


class PageCache(object):

    _cache = {}
    cacheDir = None
    _fileMapLoaded = False
    _fileMapName = 'url_file.map'
    _maxFileNumber = None
    memoryCache = True
    onlyFromCache = False

    @staticmethod
    def getPage(url, withoutCache=False, oldAttemptsPages=None):
        if oldAttemptsPages is not None and len(oldAttemptsPages) > 0:
            withoutCache=True
        url = normalizeUrl(url)
        # TODO move to normalizeUrl
        if type(url) is str:
            url = url.decode(ParsBase._encoding)
        container = PageCache._cache.get(url, None)
        if container is None:
            container = Container()
            page, fileName = PageCache._loadPage(url, withoutCache
                                                 , oldAttemptsPages)
            if PageCache.memoryCache:
                container.page = page
            else:
                container.page = None
            container.fileName = fileName
            PageCache._cache[url] = container
        else:
            if container.page is None or withoutCache:
                page, fileName = PageCache._loadPage(url, withoutCache
                                                     , oldAttemptsPages)
                if PageCache.memoryCache:
                    container.page = page
                else:
                    container.page = None
                container.fileName = fileName
            else:
                page = container.page
        return page

    @staticmethod
    def _webLoad(url, oldAttemptsPages):
        if PageCache.onlyFromCache:
            if type(url) is not str:
                if type(url) is not unicode:
                    url = unicode(url, ParsBase._encoding)
                url = url.encode(ParsBase._encoding)
            raise PageCacheWarning(url+' not in cache')
        page = Web.getGrabPage(url, oldAttemptsPages)
        return page

    @staticmethod
    def _webLoadSaveFile(url, oldAttemptsPages):
        page = PageCache._webLoad(url, oldAttemptsPages)
        fileName = PageCache._writePage(page)
        PageCache._saveFileMap(url, fileName)
        return page, fileName

    @staticmethod
    def _path(fileName):
        path = PageCache.cacheDir
        if path[len(path)-1] != '/':
            path += '/'
        path += fileName
        return path

    @staticmethod
    def _findMaxFileNumber():
        if not PageCache._fileMapLoaded:
            PageCache._loadFileMap()
        cache = PageCache._cache
        maxNumber = 0
        for url in cache:
            container = cache[url]
            if container is not None:
                fileName = container.fileName
                if fileName is not None:
                    fileNumber = int(fileName[:-7])  # remove extension .pickle
                    if maxNumber < fileNumber:
                        maxNumber = fileNumber
        return maxNumber

    @staticmethod
    def _createFileName():
        if not PageCache._fileMapLoaded:
            PageCache._loadFileMap()
        maxFileNumber = PageCache._maxFileNumber
        if maxFileNumber is None:
            maxFileNumber = PageCache._findMaxFileNumber()
        maxFileNumber += 1
        fileName = str(maxFileNumber) + '.pickle'
        PageCache._maxFileNumber = maxFileNumber
        return fileName

    @staticmethod
    def _writePage(page, fileName=None):
        if fileName is None:
            fileName = PageCache._createFileName()
        path = PageCache._path(fileName)
        pageFile = open(path, 'wb')
        pagePickler = pickle.Pickler(pageFile, pickle.HIGHEST_PROTOCOL)
        pagePickler.dump(page)
        pageFile.close()
        return fileName

    @staticmethod
    def _readPage(fileName):
        path = PageCache._path(fileName)
        try:
            pageFile = open(path, 'rb')
        except IOError as e:
            if e.errno == 2:
                return None
            else:
                raise
        try:
            pageUnpickler = pickle.Unpickler(pageFile)
            page = pageUnpickler.load()
        except Exception:
            return None
        pageFile.close()
        return page

    @staticmethod
    def _loadFileMap():
        path = PageCache._path(PageCache._fileMapName)
        try:
            mapFile = open(path, 'r')
        except IOError as e:
            if e.errno == 2:
                return
            else:
                raise
        for mapString in mapFile:
            split = mapString.strip().split('\t')
            url = split[0]
            url = normalizeUrl(url)
            fileName = split[1]
            container = Container()
            container.page = None
            container.fileName = fileName
            if type(url) is str:
                url = url.decode(ParsBase._encoding)
            PageCache._cache[url] = container

    @staticmethod
    def _saveFileMap(url, fileName):
        if type(url) is unicode:
            url = url.encode(ParsBase._encoding)
        if type(fileName) is unicode:
            fileName = fileName.encode(ParsBase._encoding)
        path = PageCache._path(PageCache._fileMapName)
        mapFile = open(path, 'a')
        mapString = '{0}\t{1}\n'.format(url, fileName)
        mapFile.write(mapString)
        mapFile.close()

    @staticmethod
    def _fileLoad(url, withoutCache=False, oldAttemptsPages=None):
        if not PageCache._fileMapLoaded:
            PageCache._loadFileMap()
            PageCache._fileMapLoaded = True
        container = PageCache._cache.get(url, None)
        if withoutCache:
            if container is None:
                page, fileName = PageCache._webLoadSaveFile(url
                                                            , oldAttemptsPages)
            elif container.fileName is None:
                page, fileName = PageCache._webLoadSaveFile(url, oldAttemptsPages)
            else:
                fileName = container.fileName
                page = PageCache._webLoad(url, oldAttemptsPages)
                fileName = PageCache._writePage(page, fileName)
        elif container is None:
            page, fileName = PageCache._webLoadSaveFile(url, oldAttemptsPages)
        elif container.page is None and container.fileName is None:
            page, fileName = PageCache._webLoadSaveFile(url, oldAttemptsPages)
        elif container.page is None:
            fileName = container.fileName
            page = PageCache._readPage(fileName)
            if page is None:
                page = PageCache._webLoad(url, oldAttemptsPages)
                fileName = PageCache._writePage(page, fileName)
        elif container.fileName is None:
            page = container.page
            fileName = PageCache._writePage(page)
            PageCache._saveFileMap(url, fileName)
        else:
            page = container.page
            fileName = container.fileName
        return page, fileName

    # @staticmethod
    # def _mkdir(dirPath):
        # path = PageCache.cacheDir
        # if os.path.exists(path):
            # if not os.path.isdir(path):
                # raise PageCacheError('Cannot create dir: '+path)
        # else:
            # os.makedirs(path)

    @staticmethod
    def _loadPage(url, withoutCache=False, oldAttemptsPages=None):
        if PageCache.cacheDir is None:
            fileName = None
            page = PageCache._webLoad(url, oldAttemptsPages)
        else:
            # PageCache._mkdir(PageCache.cacheDir)
            mkdirs(PageCache.cacheDir)
            page, fileName = PageCache._fileLoad(url, withoutCache
                                                 , oldAttemptsPages)
        return page, fileName

    @staticmethod
    def _proxyType():
        regex = u'/proxy-([^.]+)\.list$'
        pattern = RegexCache.compile(regex)
        proxyType = pattern.findall(PageCache.proxyFile.strip())[0]
        if proxyType in ('http', 'socks4', 'socks5'):
            return proxyType
        else:
            raise PageCacheError('proxy server type: ' + proxyType\
                                 + ' is invalid')

    @staticmethod
    def getPageFromCache(url):
        url = normalizeUrl(url)
        if type(url) is str:
            url = url.decode(ParsBase._encoding)
        container = PageCache._cache.get(url, None)
        if container is None:
            return None
        page = None
        page = container.page
        if page is None and container.fileName is not None:
            page = PageCache._readPage(container.fileName)
            if PageCache.memoryCache and page is not None:
                container.page = page
        return page

    @staticmethod
    def rewritePageInDisk(url):
        if PageCache.cacheDir is None:
            return
        url = normalizeUrl(url)
        if type(url) is str:
            url = url.decode(ParsBase._encoding)
        print(url.encode(ParsBase._encoding))
        container = PageCache._cache.get(url, None)
        if container is None:
            return
        if container.page is None:
            return
        fileName = PageCache._writePage(container.page, container.fileName)
        if fileName != container.fileName:
            PageCache._saveFileMap(url, fileName)
            container.fileName = fileName

    @staticmethod
    def writePageInCache(url, page):
        url = normalizeUrl(url)
        if type(url) is str:
            url = url.decode(ParsBase._encoding)
        if PageCache.cacheDir is not None:
            if not PageCache._fileMapLoaded:
                PageCache._loadFileMap()
                PageCache._fileMapLoaded = True
        container = PageCache._cache.get(url, None)
        if container is None:
            container = Container()
            container.fileName = None
            container.page = None
        if PageCache.cacheDir is not None:
            fileName = PageCache._writePage(page, container.fileName)
            if fileName != container.fileName:
                PageCache._saveFileMap(url, fileName)
                container.fileName = fileName
        else:
            container.fileName = None
        if PageCache.memoryCache:
            container.page = page
        else:
            container.page = None
        if container.page is not None or container.fileName is not None:
            PageCache._cache[url] = container
        else:
            _ = PageCache._cache.pop(url)


class PageBase(ParsBase):

    attempts = 5
    errorWait = 60
    urlControl = False

    def _old__init__(self, *args, **kwargs):
        self._urlControl = kwargs.pop('urlControl', None)
        needControl = kwargs.pop('needControl', True)
        ParsBase.__init__(self, *args, **kwargs)
        self._needControl = needControl

    def _old_processing(self):
        self._url = normalizeUrl(self._elem)
        i = 0
        withoutCache = False
        urlControl = self._urlControl
        if urlControl is None:
            urlControl = Page.urlControl
        selfUrl = normalizeUrl(self._url)
        while i < Page.attempts:
            i += 1
            page = PageCache.getPage(selfUrl, withoutCache)
            pageUrl = getattr(page.response, 'url', selfUrl)
            pageUrl = normalizeUrl(pageUrl)
            if urlControl and selfUrl.lower() != pageUrl.lower():
                if i >= Page.attempts:
                    msg = 'Fails urlControl'
                    raise PageError(msg, grabPage=page)
                else:
                    withoutCache = True
                    time.sleep(Page.errorWait)
            else:
                self._elem = page.xpath('/*')
                return
        msg = 'Fails urlControl'
        raise PageError(msg, grabPage=page)

    def __init__(self, *args, **kwargs):
        self._page = None
        needControl = kwargs.pop('needControl', True)
        ParsBase.__init__(self, *args, **kwargs)
        self._needControl = needControl

    def _processing(self, oldAttemptsInstances=None):
        self._url = normalizeUrl(self._elem)
        if oldAttemptsInstances is None or oldAttemptsInstances == []:
            page = PageCache.getPage(self._url, withoutCache=False)
        else:
            oldAttemptsPages = []
            for oldInstance in oldAttemptsInstances:
                if oldInstance._page is not None:
                    oldAttemptsPages.append(oldInstance._page)
            page = PageCache.getPage(self._url, withoutCache=False
                                     , oldAttemptsPages=oldAttemptsPages)
        if self._needControl:
            if page.pageConfirmed:
                self._needControl = False
            else:
                self._needControl = True
        self._page = page
        # if not self._needControl:
            # # not working!!! extra entries in the log
            # self._page.regProxyGoodPage()
        self._elem = None

    def _regEventPagesProxy(self, oldAttemptsInstances):
        if oldAttemptsInstances is None:
            return
        identicalInstances = self._findIdenticalInstances(oldAttemptsInstances)
        for instance in oldAttemptsInstances:
            if instance in identicalInstances:
                instance._page.regProxyGoodPage()
            else:
                instance._page.regProxyBadPage()

    def _instanceControl(self, oldAttemptsInstances=None):
        # print('enterContro')
        if self._page.pageConfirmed:
            self._regEventPagesProxy(oldAttemptsInstances)
            return ControlResult.ok
        cachePage = PageCache.getPageFromCache(self._url)
        if cachePage is not None:
            if cachePage.pageConfirmed:
                if cachePage.uuid == self._page.uuid:
                    self._regEventPagesProxy(oldAttemptsInstances)
                    return ControlResult.ok
                else:
                    return ControlResult.bad  # The next time will be obtained from the cache page
        controlResult = ParsBase._instanceControl(self, oldAttemptsInstances)
        if controlResult == ControlResult.ok:
            self._page.pageConfirmed = True
            self._page.regProxyGoodPage()
            if cachePage is None:
                PageCache.writePageInCache(self._url, self._page)
            elif cachePage.uuid == self._page.uuid:
                if not cachePage.pageConfirmed:
                    cachePage.pageConfirmed = True
                    # PageCache.rewritePageInDisk(self._url)
                    PageCache.writePageInCache(self._url, cachePage)
            else:
                PageCache.writePageInCache(self._url, self._page)
            self._regEventPagesProxy(oldAttemptsInstances)
        # print(controlResult)
        return controlResult

    def _calcHash(self):
        value = self._url
        if self._page.page.response.code == 404:
            value = '404'
        if type(value) is unicode:
            value = value.encode(ParsBase._encoding)
        return md5(value)


class Page(XpathQueryMixin, UnicodeTreeMixin, PageBase):

    def _processing(self, oldAttemptsInstances=None):
        PageBase._processing(self, oldAttemptsInstances)
        if self._page.page.response.code != 404:
            self._elem = self._page.page.xpath('/*')
        else:
            self._elem = None


class File(PageBase):

    def __init__(self, *args, **kwargs):
        self._file = None
        PageBase.__init__(self, *args, **kwargs)
        self._homeDir = kwargs.get('homeDir', u'./')
        self._homeDir = unicode(self._homeDir, ParsBase._encoding)
        self._homeDir = normalizePath(self._homeDir, itDir=True)
        self._dirForFile = kwargs.get('dirForFile', u'')
        self._dirForFile = unicode(self._dirForFile, ParsBase._encoding)
        self._dirForFile = normalizePath(self._dirForFile, itDir=True)
        self._showHomeDir = kwargs.get('showHomeDir', True)
        self._filePathType = kwargs.get('filePathType', FilePathType.urlPath)
        if self._filePathType not in (FilePathType.urlPath
                                     , FilePathType.uuidSingleDir
                                     , FilePathType.uuidMultiDir):
            raise ParsError('filePathType must be in FilePathType.urlPath,' \
                            + ' FilePathType.uuidSingleDir,' \
                            + ' FilePathType.uuidMultiDir')

    def _createFilePath(self):
        urlPath = self._url_path
        if type(urlPath) is not unicode:
            urlPath = unicode(urlPath, ParsBase._encoding)
        if self._filePathType == FilePathType.urlPath:
            return urlPath
        _, orignalFileName = splitDirFile(urlPath)
        extension = orignalFileName.rsplit(u'.', 1)[1]
        fileName = uuid.uuid4().hex
        if len(extension) > 0:
            fileName += u'.' + extension
        if self._filePathType == FilePathType.uuidSingleDir:
            return fileName
        elif self._filePathType == FilePathType.uuidMultiDir:
            path = fileName[0:2] + u'/' + fileName[2:4] + u'/' + fileName
            return path
        else:
            raise ParsError('filePathType must be in FilePathType.urlPath,' \
                            + ' FilePathType.uuidSingleDir,' \
                            + ' FilePathType.uuidMultiDir')

    def _processing(self, oldAttemptsInstances=None):
        PageBase._processing(self, oldAttemptsInstances)
        if self._page.page.response.code != 404:
            self._elem = self._page.page.response.body
        else:
            self._elem = None
        self._localPath = self._createFilePath()

    def _write(self):
        path = self._homeDir + self._dirForFile + self._localPath
        if self._elem is not None:
            write(self._elem, path, 'wb')

    @property
    def _unicode(self):
        if self._page.page.response.code == 404:
            return u''
        path = self._dirForFile + self._localPath
        if self._showHomeDir:
            path = self._homeDir + path
        return normalizePath(path)

    def _calcHash(self):
        if self._page.page.response.code == 404:
            return md5('404')
        return md5(self._elem)


class PagerMixin(object):

    def _listInstanceConstruct(self, queryResult, saveInstance=False):
        saveInstance = saveInstance or self._saveInstance
        queryResult = queryResult[0]
        url = queryResult['startUrl']
        href = queryResult['href']
        resultList = []
        iterNumber = 0
        while True:
            if self._maxIteration is not None:
                iterNumber += 1
                if iterNumber > self._maxIteration:
                    break
            page = self._instanceConstruct(url, saveInstance=True)
            if saveInstance:
                resultList.append(page)
            if self._breakPage(page):
                break
            url = page._href(href)
            if len(url) == 0:
                break
            if len(url) > 1:
                raise BadQueryResult('Query "href" returned more than one value')
            url = url[0]
        if saveInstance:
            return resultList
        else:
            return None

    def _breakPage(self, page):
        return False


class Pager(PagerMixin, Page):
    pass


class KeyStr(Str):

    def _processing(self):
        Str._processing(self)
        self._elem = normalizeSpace(self._elem)


class KeyText(Text):

    def _processing(self):
        Text._processing(self)
        self._elem = normalizeSpace(self._elem)


# -------------------------------------------------------------------------

@staticmethod
def urlControl(page, url, proxy):
    response = getattr(page, 'response', None)
    if response is None:
        return
    #----------------------
    # reqUrl = url
    # resUrl = getattr(response, 'url', None)
    # print('type reqUrl =', type(reqUrl))
    # print('type resUrl =', type(resUrl))
    # print('reqUrl =', reqUrl.encode(ParsBase._encoding))
    # print('resUrl =', resUrl)
    # nResUrl = normalizeUrl(resUrl)
    # print('type nResUrl =', type(nResUrl))
    # print('nResUrl =', nResUrl)
    # nResUrl = nResUrl.decode(ParsBase._encoding)
    # print('type nResUrl =', type(nResUrl))
    # print('nResUrl =', nResUrl.encode(ParsBase._encoding))
    # sys.exit(0)
    #----------------------
    requestUrl = normalizeUrl(url)
    responseUrl = normalizeUrl(getattr(response, 'url', requestUrl))
    if type(requestUrl) is str:
        requestUrl = requestUrl.decode(ParsBase._encoding)
    if type(responseUrl) is str:
        responseUrl = responseUrl.decode(ParsBase._encoding)
    if requestUrl.lower() != responseUrl.lower():
        httpCode = getattr(response, 'code', None)
        httpBody = getattr(response, 'body', None)
        raise PageControlFault(httpCode=httpCode, httpBody=httpBody
                               , requestUrl=requestUrl, responseUrl=responseUrl)


class PrintTreeCatcher(object):

    def __init__(self, fileName=None):
        self.fileName = fileName

    def __call__(self, instance):
        instance.printTree(fileName=self.fileName)


class MultiCatcher(object):

    def __init__(self, *args):
        self.catchers = args

    def __call__(self, instance):
        for catcher in self.catchers:
            catcher(instance)


class BaseHashControlCatcher(object):

    def __init__(self, **kwargs):
        self.fileDir = kwargs.get('fileDir')
        self.stopCnt = kwargs.get('stopCnt', None)
        self.continuousRepeats = kwargs.get('continuousRepeats', True)
        self.useSessions = kwargs.get('useSessions', 1)
        self.progName = kwargs.get('progName', None)
        self.extension = kwargs.get('extension', 'treehash')
        self.saveSessionCnt = kwargs.get('saveSessionCnt', 1)
        self.hashSet = None
        self.duplicateCnt = 0
        self.fileName = None

    def loadHashList(self):
        self.hashSet = set()
        pattern = self.fileDir + '*.'
        if self.progName is not None:
            pattern += self.progName + '.'
        pattern += self.extension
        fileSet = set(glob.glob(pattern))
        i = 0
        for fileName in sorted(fileSet, reverse=True):
            i += 1
            if i > self.useSessions:
                break
            f = open(fileName, 'r')
            for line in f:
                self.hashSet.add(line.rstrip())
            f.close()
        if self.stopCnt > len(self.hashSet) and len(self.hashSet) > 0:
            self.stopCnt = len(self.hashSet)
        if len(fileSet) > self.saveSessionCnt:
            delCnt = len(fileSet) - self.saveSessionCnt
            i = 0
            for fileName in sorted(fileSet):
                i += 1
                if i > delCnt:
                    break
                os.remove(fileName)
                fileSet.remove(fileName)

    def saveHash(self, hash):
        fileName = self.fileName
        if fileName is None:
            mkdirs(self.fileDir)
            fileName = self.fileDir + createDatePidFileName()
            if self.progName is not None:
                fileName += '.' + self.progName
            fileName += '.' + self.extension
            self.fileName = fileName
        f = open(fileName, 'a')
        f.write(hash + '\n')
        f.close()

    def calcHash(self, instance):
        raise NotImplementedError

    def __call__(self, instance):
        if self.hashSet is None:
            self.loadHashList()
        curHash = self.calcHash(instance)
        if curHash in self.hashSet:
            self.duplicateCnt += 1
            if self.stopCnt is not None and self.stopCnt > 0:
                if self.duplicateCnt >= self.stopCnt:
                    raise StopParsing()
            raise DuplicateTree()
        else:
            self.saveHash(curHash)
            if self.continuousRepeats:
                self.duplicateCnt = 0


class TreeHashControlCatcher(BaseHashControlCatcher):

    def calcHash(self, instance):
        return instance._getTreeHash()


class MultiHashConrolCatcher(BaseHashControlCatcher):

    def __init__(self, *args, **kwargs):
        self.controls = args
        BaseHashControlCatcher.__init__(self, **kwargs)

    def calcHash(self, instance):
        hash = ''
        for control in self.controls:
            obj = control(instance)
            hash += getHash(obj)
        return md5(hash)


class TimeControlCatcher(object):

    def __init__(self, fileDir , timeCatcher, progName=None, saveSessionCnt=1):
        self.fileDir = fileDir
        self.progName = progName
        self.timeCatcher = timeCatcher
        self.saveSessionCnt = saveSessionCnt
        self.useSessions = 1
        self.oldTime = None
        self.curTime = datetime.datetime(1970, 1, 1)
        self.fileName = None

    def loadOldTime(self):
        self.oldTime = datetime.datetime(1970, 1, 1)
        pattern = self.fileDir + '*.'
        if self.progName is not None:
            pattern += self.progName + '.'
        pattern += 'oldtime'
        fileSet = set(glob.glob(pattern))
        # --------
        # print(fileSet)
        # for fileName in sorted(fileSet):
            # print(fileName)
        # print('---------------------')
        # ----------------
        i = 0
        for fileName in sorted(fileSet, reverse=True):
            i += 1
            if i > self.useSessions:
                break
            f = open(fileName, 'r')
            for line in f:
                line = line.rstrip()
                dt = datetime.datetime.strptime(line, '%Y-%m-%d-%H-%M-%S')
                if dt > self.oldTime:
                    self.oldTime = dt
            f.close()
        if len(fileSet) > self.saveSessionCnt:
            delCnt = len(fileSet) - self.saveSessionCnt
            i = 0
            for fileName in sorted(fileSet):
                i += 1
                if i > delCnt:
                    break
                os.remove(fileName)
                fileSet.remove(fileName)

    def saveTime(self, dt):
        fileName = self.fileName
        if fileName is None:
            mkdirs(self.fileDir)
            fileName = self.fileDir + createDatePidFileName()
            if self.progName is not None:
                fileName += '.' + self.progName
            fileName += '.oldtime'
            self.fileName = fileName
        dt = dt.strftime('%Y-%m-%d-%H-%M-%S')
        f = open(fileName, 'a')
        f.write(dt + '\n')
        f.close()

    def __call__(self, instance):
        if self.oldTime is None:
            self.loadOldTime()
        curTime = self.timeCatcher(instance)
        if not isinstance(curTime, datetime.datetime):
            return
        if self.oldTime is not None and curTime < self.oldTime:
            raise StopParsing()
        else:
            if curTime > self.curTime and curTime > self.oldTime:
                self.saveTime(curTime)
                self.curTime = curTime

# -----------------------------------------------------------------------

import openpyxl


class XlsxCell(object):

    def __init__(self, sheet):
        self.sheet = sheet
        self.catcher = XlsxCell.defaultCather

    @staticmethod
    def defaultCather(instance):
        return None

    def __getattr__(self, name):
        if name == 'defaultWhenException':
            return self.sheet.defaultWhenException
        elif name == 'default':
            return self.sheet.default

    def runCatcher(self, instance):
        try:
            return self.catcher(instance)
        except Exception:
            if self.defaultWhenException:
                return self.default
            else:
                raise


class XlsxSheet(object):

    disableFormulaDefault = False

    def __init__(self, book, workSheet):
        self.book = book
        self.ws = workSheet
        self.currentRow = 1
        self.header = {}
        self.cell = {}

    def __getattr__(self, name):
        if name == 'defaultWhenException':
            return self.book.defaultWhenException
        if name == 'disableFormula':
            return XlsxSheet.disableFormulaDefault
        elif name == 'default':
            return self.book.default

    def __getitem__(self, cellIndex):
        cellIndex = self.header.get(cellIndex, cellIndex)
        # TODO Add cellIndex check the correctness of
        cell = self.cell.get(cellIndex, None)
        if cell is None:
            cell = XlsxCell(self)
            self.cell[cellIndex] = cell
        return cell

    def catcher(self, instance):
        for columnName in self.cell:
            value = self.cell[columnName].runCatcher(instance)
            # ------ disableFormula ----------
            if self.disableFormula:
                disableFormula = False
                if type(value) is str:
                    disableFormula = True
                    formulaSymbol = '='
                if type(value) is unicode:
                    disableFormula = True
                    formulaSymbol = u'='
                if disableFormula:
                    position = 0
                    findFormula = False
                    while position < len(value):
                        ch = value[position]
                        if ch == formulaSymbol:
                            findFormula = True
                            position += 1
                        elif ch.isspace():
                            position += 1
                        else:
                            break
                    if findFormula:
                        value = value[position:]
            # ------ end disableFormula ------
            rowName = columnName + str(self.currentRow)
            value = self.decoding(value)
            self.ws[rowName] = value
        self.currentRow += 1

    def decoding(self, value):
        if type(value) is str:
            value = value.decode(ParsBase._encoding)
        if type(value) is unicode:
            value = value.encode(self.book.encoding)
        return value

    def writeHeader(self):
        if len(self.header) == 0:
            return
        for head in self.header:
            columnName = self.header[head]
            columnName = self.decoding(columnName)
            rowName = columnName + str(self.currentRow)
            self.ws[rowName] = head
        self.currentRow += 1


class XlsxError(Exception):
    pass


class XlsxCatcher(object):

    def __init__(self, fileName=None
                 , optimized_write=False
                 , encoding='utf-8'
                 , guess_types=False
                 , data_only=False
                 , read_only=False
                 , write_only=False):
        self.encoding = encoding
        self.fileName = fileName
        self.sheet = {}
        self.sheetList = []
        self.defaultWhenException = False
        self.wb = openpyxl.Workbook(
            optimized_write=optimized_write
            , encoding=encoding
            , guess_types=guess_types
            , data_only=data_only
            , read_only=read_only
            , write_only=write_only
        )

    def __getitem__(self, sheetIndex):
        if type(sheetIndex) is int:
            sheetIndex = self.sheetList[sheetIndex]
        if type(sheetIndex) is unicode:
            sheetIndex = sheetIndex.encode(self.encoding)
        sheet = self.sheet.get(sheetIndex, None)
        if sheet is None:
            if len(self.sheet) == 0:
                ws = self.wb.active
            else:
                ws = self.wb.create_sheet()
            ws.title = sheetIndex
            sheet = XlsxSheet(self, ws)
            self.sheet[sheetIndex] = sheet
            self.sheetList.append(sheetIndex)
        return sheet

    def catcher(self, instance):
        for sheetName in self.sheet:
            self.sheet[sheetName].catcher(instance)

    def __call__(self, instance):
        self.catcher(instance)

    def save(self, fileName=None):
        if fileName is None:
            fileName = self.fileName
        if fileName is None:
            raise XlsxError('undefined fileName')
        dirs, _ = splitDirFile(fileName)
        mkdirs(dirs)
        self.wb.save(fileName)

    def writeHeader(self):
        for sheetName in self.sheet:
            self.sheet[sheetName].writeHeader()


class MaxTransactionsControlCatcher(object):

    def __init__(self, maxTransactions):
        self.maxTransactions = maxTransactions
        self.transactionNumber = 0

    def __call__(self, instance):
        self.transactionNumber += 1
        if self.transactionNumber > self.maxTransactions:
            raise StopParsing


# ----------------------------------------------------------------------

def list2unicode(parsList, divider=u' '):
    result = u''
    if type(divider) is str:
        divider = divider.encode(ParsBase._encoding)
    for elem in parsList:
        if elem is None or unicode(elem) == u'':
            continue
        uniStr = unicode(elem)
        uniStr = uniStr.strip()
        # if uniStr[-1] != u'.':
            # uniStr += u'.'
        # if len(result) > 0 and result[-1] != u'\n':
        if len(result) > 0:
            uniStr = divider + uniStr
        result += uniStr
    return result


if __name__ == '__main__':
    pass
